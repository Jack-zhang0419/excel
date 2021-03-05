from copy import copy
from openpyxl.worksheet.worksheet import Worksheet


# https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/worksheet/copier.html
# Copy range of cells as a nested list
# Takes: start cell, end cell, and sheet you want to copy from.
def copy_range(start_column: int, start_row: int, end_column: int,
               end_row: int, source_sheet: Worksheet):
    range_selected = []
    # Loops through selected Rows
    for i in range(start_row, end_row + 1, 1):
        # Appends the row to a row_selected list
        row_selected = []
        for j in range(start_column, end_column + 1, 1):
            row_selected.append(source_sheet.cell(row=i, column=j))
        # Adds the row_selected List and nests inside the rangeSelected
        range_selected.append(row_selected)

    return range_selected


# Paste range to target worksheet include data and styles
def paste_range(start_column: int, start_row: int, end_column: int,
                end_row: int, target_cells, source_cells):
    count_row = 0

    for i in range(start_row, end_row + 1, 1):
        count_column = 0
        for j in range(start_column, end_column + 1, 1):
            target_cell = target_cells.cell(row=i, column=j)
            source_cell = source_cells[count_row][count_column]

            target_cell.data_type = source_cell.data_type
            target_cell.value = source_cell.value

            if source_cell.has_style:
                # commented because probably the style include operating system related properties, such as '常规' font data_type
                # target_cell.style = source_cell.style
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)

            if source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)

            if source_cell.comment:
                target_cell.comment = copy(source_cell.comment)

            # thin = Side(border_style="thin", color="000000")
            # double = Side(border_style="double", color="ff0000")

            # target_cell.border = Border(top=thin,
            #                             left=thin,
            #                             right=thin,
            #                             bottom=thin)
            # target_cell.fill = PatternFill("solid", fgColor="DDDDDD")
            # target_cell.fill = fill = GradientFill(stop=("000000",
            #                                              "FFFFFF"))
            # target_cell.font = Font(b=True, color="FF0000")
            # target_cell.alignment = Alignment(horizontal="center",
            #                                   vertical="center")

            count_column += 1
        count_row += 1
