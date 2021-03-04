import openpyxl
from openpyxl.worksheet.cell_range import CellRange
from task.itask import ITask
from util.file_util import sub_dirs, filter_excel_files
from openpyxl import load_workbook, workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.copier import WorksheetCopy
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from copy import copy
import os

# https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/worksheet/cell_range.html?highlight=CellRange#
SHEET_NAME = ["学校层面", "专业群层面"]


class MergeAB(ITask):
    """
    merge multiple excels into one according with file name
    """
    output_file = "merged.xlsx"

    def run(self):
        print(
            '================================================================')
        print(f"start to merge '{self.dir}' folder")

        root_folder = f"{os.getcwd()}/{self.dir}"

        self.merge(root_folder)

    def merge(self, folder):
        """
        docstring
        """
        print(
            '----------------------------------------------------------------')
        print(f'start to merge folder: {folder} ......')
        # 文件路径
        # 构建新的表格名称
        merged_file = f"{folder}/{self.output_file}"
        self.target_workbook = workbook.Workbook()
        self.target_workbook.active.title = SHEET_NAME[0]
        self.target_workbook.create_sheet(title=SHEET_NAME[1])
        self.target_cur_row = [1, 1]

        filter_list = filter_excel_files(folder)
        if self.output_file in filter_list:
            filter_list.remove(self.output_file)
        sorted_list = sorted(filter_list)

        for file_name in sorted_list:
            source_sheet_no, source_block_no, target_sheet_no, target_block_no = self.parse_excel_name(
                file_name)
            self.source_workbook = load_workbook(
                filename=os.path.join(folder, file_name))

            (copied_range, column_range,
             source_row_range) = self.copy_excel_block(source_sheet_no,
                                                       source_block_no)
            current_last_row = source_row_range[1] - source_row_range[
                0] + self.target_cur_row[target_sheet_no]

            self.paste_range(1, self.target_cur_row[target_sheet_no],
                             column_range, current_last_row,
                             self.target_workbook.worksheets[target_sheet_no],
                             copied_range)

            # handle worksheet level
            self.target_workbook.worksheets[
                target_sheet_no].sheet_format = copy(
                    self.source_workbook.worksheets[source_sheet_no].
                    sheet_format)
            self.target_workbook.worksheets[
                target_sheet_no].sheet_properties = copy(
                    self.source_workbook.worksheets[source_sheet_no].
                    sheet_properties)
            for attr in ('row_dimensions', 'column_dimensions'):
                src = getattr(self.source_workbook.worksheets[source_sheet_no],
                              attr)
                target = getattr(
                    self.target_workbook.worksheets[target_sheet_no], attr)
                for key, dim in src.items():
                    target[key] = copy(dim)
                    target[key].worksheet = self.target_workbook.worksheets[
                        target_sheet_no]

            # handle merged_cells
            source_block_area = CellRange(
                min_col=1,
                max_col=column_range,
                min_row=self.target_cur_row[target_sheet_no],
                max_row=source_row_range[1] - source_row_range[0] +
                self.target_cur_row[target_sheet_no])

            for mcr in self.source_workbook.worksheets[
                    source_sheet_no].merged_cell_ranges:
                if mcr.coord in source_block_area:
                    cr = CellRange(mcr.coord)
                    cr.shift(row_shift=self.target_cur_row[target_sheet_no] -
                             source_row_range[0])
                    self.target_workbook.worksheets[
                        target_sheet_no].merge_cells(cr.coord)

            self.target_cur_row[target_sheet_no] = current_last_row + 1
            self.source_workbook = None
        self.target_workbook.save(merged_file)

    def parse_excel_name(self, file_name):
        current_file_name = file_name.split(".")[0]
        if current_file_name[0] == "A":
            source_sheet_no = target_sheet_no = 0
        else:
            source_sheet_no = target_sheet_no = 1
        source_block_no = target_block_no = int(current_file_name[1])

        if "-" in current_file_name:
            source_block_no = int(current_file_name[-1])
            if current_file_name[-2] == "A":
                source_sheet_no = 0
            else:
                source_sheet_no = 1

        return (source_sheet_no, source_block_no, target_sheet_no,
                target_block_no)

    def get_row_range(self, sheet: Worksheet, block_no: int):
        bounds = sorted([
            merged_range.bounds for merged_range in sheet.merged_cells.ranges
            if merged_range.bounds[0] == 1
        ])
        first_row_block = bounds[block_no]
        if block_no == 1:
            found = (1, first_row_block[3])
        else:
            found = (first_row_block[1], first_row_block[3])
        print(f"block {block_no} row range: {found[0]} to {found[1]}")
        return found

    def get_column_range(self, sheet: Worksheet):
        column = 1
        while type(sheet.cell(
                row=1, column=column)).__name__ == 'MergedCell' or sheet.cell(
                    row=1, column=column).value is not None:
            column = column + 1

        print(f'current sheet {sheet.title} max column: {column - 1}')
        return column - 1

    def copy_excel_block(self, source_sheet_no, source_block_no):

        column_range = self.get_column_range(
            self.source_workbook.worksheets[source_sheet_no])
        row_range = self.get_row_range(
            self.source_workbook.worksheets[source_sheet_no], source_block_no)

        return (self.copy_range(
            1, row_range[0], column_range, row_range[1],
            self.source_workbook.worksheets[source_sheet_no]), column_range,
                row_range)

    # Copy range of cells as a nested list
    # Takes: start cell, end cell, and sheet you want to copy from.
    # https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/worksheet/copier.html
    def copy_range(self, startCol, startRow, endCol, endRow, sheet):
        rangeSelected = []
        # Loops through selected Rows
        for i in range(startRow, endRow + 1, 1):
            # Appends the row to a RowSelected list
            rowSelected = []
            for j in range(startCol, endCol + 1, 1):
                rowSelected.append(sheet.cell(row=i, column=j))
            # Adds the RowSelected List and nests inside the rangeSelected
            rangeSelected.append(rowSelected)

        return rangeSelected

    # Paste range
    # Paste data from copyRange into template sheet
    def paste_range(self, startCol, startRow, endCol, endRow, target_cells,
                    source_cells):
        countRow = 0

        for i in range(startRow, endRow + 1, 1):
            countCol = 0
            for j in range(startCol, endCol + 1, 1):
                target_cell = target_cells.cell(row=i, column=j)
                source_cell = source_cells[countRow][countCol]

                target_cell.data_type = source_cell.data_type
                target_cell.value = source_cell.value

                if source_cell.has_style:
                    # target_cell.style = source_cell.style
                    target_cell.font = copy(source_cell.font)
                    target_cell.border = copy(source_cell.border)
                    target_cell.fill = copy(source_cell.fill)
                    target_cell.number_format = copy(source_cell.number_format)
                    target_cell.protection = copy(source_cell.protection)
                    target_cell.alignment = copy(source_cell.alignment)

                if source_cell.hyperlink:
                    target_cell._hyperlink = copy(source_cell.hyperlink)

                if source_cell.comment:
                    target_cell.comment = copy(source_cell.comment)

                # thin = Side(border_style="thin", color="000000")
                # double = Side(border_style="double", color="ff0000")

                # target_cell.border = Border(top=thin,
                #                             left=thin,
                #                             right=thin,
                #                             bottom=thin)
                # target_cell.fill = PatternFill("solid", fgColor="DDDDDD")
                # target_cell.fill = fill = GradientFill(stop=("000000",
                #                                              "FFFFFF"))
                # target_cell.font = Font(b=True, color="FF0000")
                # target_cell.alignment = Alignment(horizontal="center",
                #                                   vertical="center")

                countCol += 1
            countRow += 1
