from copy import copy
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
from openpyxl.styles import PatternFill
from task.combine_util.configure import *
from datetime import datetime, timedelta


# https://openpyxl.readthedocs.io/en/stable/styles.html
# https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/worksheet/copier.html
# Copy range of cells as a nested list
# Takes: start cell, end cell, and sheet you want to copy from.
def copy_range(start_column: int, start_row: int, end_column: int,
               end_row: int, source_sheet: Worksheet):
    range_selected = []
    # Loops through selected Rows
    for i in range(start_row, end_row + 1, 1):
        # Appends the row to a row_selected list
        row_selected = []
        for j in range(start_column, end_column + 1, 1):
            row_selected.append(source_sheet.cell(row=i, column=j))
        # Adds the row_selected List and nests inside the rangeSelected
        range_selected.append(row_selected)

    return range_selected


def apply_column_type(target: Cell, source: Cell, column_type):
    if column_type["source_data_type"] == source.data_type and type(
            source.value).__name__ == column_type["source_value_type"]:
        if column_type["target_data_type"] == "d":
            converted_value = from_excel_ordinal(int(source.value))
            print(
                f"trying to convert column value {source.value} to date: {converted_value}"
            )
            target.data_type = 's'
            target._value = converted_value.strftime("yyyy/mm/dd")
            if "target_number_format" in column_type:
                target._value = converted_value.strftime(
                    column_type["target_number_format"])


# Paste range to target worksheet include data and styles
def paste_range(start_column: int, start_row: int, end_column: int,
                end_row: int, target_cells, source_cells, block_no):
    count_row = 0

    for i in range(start_row, end_row + 1, 1):
        count_column = 0
        for j in range(start_column, end_column + 1, 1):
            target_cell = target_cells.cell(row=i, column=j)
            source_cell = source_cells[count_row][count_column]

            target_cell.number_format = source_cell.number_format
            target_cell.data_type = source_cell.data_type
            target_cell.value = source_cell.value

            if source_cell.value is not None:
                column_type = defined_column_type(j)
                if column_type:
                    try:
                        apply_column_type(target_cell, source_cell,
                                          column_type)
                    except Exception as error:
                        print(
                            f"warning: cannot convert cell to column type: {type(error)} {error}"
                        )

            if source_cell.has_style:
                # commented because probably the style include operating system related properties, such as '??????' font data_type
                # target_cell.style = source_cell.style
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)

                # if master excel, then take the style of header of master
                # if BLOCK_COLORS set, then take BLOCK_COLORS
                if block_no == 0:
                    target_cell.fill = copy(source_cell.fill)
                    if HEADER_COLOR:
                        target_cell.fill = PatternFill("solid",
                                                       fgColor=HEADER_COLOR)
                if block_no != 0 and len(BLOCK_COLORS) > 0:
                    block_color_index = block_no % len(BLOCK_COLORS)
                    target_cell.fill = PatternFill(
                        "solid", fgColor=BLOCK_COLORS[block_color_index])

            if source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)

            if source_cell.comment:
                target_cell.comment = copy(source_cell.comment)

            count_column += 1
        count_row += 1


def from_excel_ordinal(ordinal, _epoch0=datetime(1899, 12, 31)):
    if ordinal >= 60:
        ordinal -= 1  # Excel leap year bug, 1900 is not a leap year!
    return (_epoch0 + timedelta(days=ordinal)).replace(microsecond=0)
