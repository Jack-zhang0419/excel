from copy import copy
from openpyxl import load_workbook, workbook
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.merge import MergedCellRange
from task.combine_util.excel_util import paste_range


class TargetExcel(object):
    """
    TargetExcel is the excel going to generate with the copied data from SourceExcel
    """
    def __init__(self, saved_file_path):
        self.saved_file_path = saved_file_path

        self.workbook = workbook.Workbook()

        self.sheet_no = 0
        self.start_column = 1  # all the start_column is 1
        self.start_rows = [1]
        self.block_nos = [0]
        self.worksheet_level_set = [False]

    def save(self):
        self.workbook.save(self.saved_file_path)

    def set_start_row(self):
        self.start_rows[self.sheet_no] = self.end_row + 1

    def increase_block_no(self):
        self.block_nos[self.sheet_no] += 1

    def switch_sheet(self, sheet_no: int):
        self.sheet_no = sheet_no
        while len(self.workbook.worksheets) <= self.sheet_no:
            self.workbook.create_sheet()
            self.start_rows.append(1)
            self.block_nos.append(1)
            self.worksheet_level_set.append(False)
        self.worksheet = self.workbook.worksheets[self.sheet_no]

    def paste_excel_block(self, start_column, start_row, end_column, end_row,
                          copied_range):
        self.end_row = self.start_rows[self.sheet_no] + end_row - start_row
        paste_range(start_column, self.start_rows[self.sheet_no], end_column,
                    self.end_row, self.worksheet, copied_range,
                    self.block_nos[self.sheet_no])
        print(
            f"target block {self.block_nos[self.sheet_no]} row range: A{self.start_rows[self.sheet_no]}:A{self.end_row}"
        )

    # https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/worksheet/cell_range.html?highlight=CellRange#
    def paste_merged_cell_range(self, source_start_row: int,
                                source_block_area: CellRange,
                                merged_cell_ranges: list[MergedCellRange]):
        for mcr in merged_cell_ranges:
            if mcr.coord in source_block_area:
                cr = CellRange(mcr.coord)
                cr.shift(row_shift=self.start_rows[self.sheet_no] -
                         source_start_row)
                self.worksheet.merge_cells(cr.coord)

    def set_worksheet_column_dimensions(self, src):
        # https://openpyxl.readthedocs.io/en/stable/api/openpyxl.worksheet.dimensions.html#openpyxl.worksheet.dimensions.ColumnDimension.width

        target = getattr(self.worksheet, 'column_dimensions')
        for key, dim in src.items():
            target[key].width = dim.width

    def set_row_dimensions(self, src, source_start_row):
        target = self.worksheet.row_dimensions
        target_start_row = self.start_rows[self.sheet_no]
        target_end_row = self.end_row

        for i in range(target_start_row, target_end_row + 1, 1):
            target[i].height = src[source_start_row].height
            source_start_row += 1
