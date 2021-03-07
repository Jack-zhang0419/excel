import os

from openpyxl import load_workbook, workbook
from openpyxl.worksheet.cell_range import CellRange
from task.combine_util.excel_util import copy_range


class SourceExcel(object):
    """docstring for SourceExcel."""
    def __init__(self, folder, excel_file, sheet_no, block_no):
        self.folder = folder
        self.excel_file = excel_file
        self.sheet_no = sheet_no
        self.block_no = block_no

        self.workbook = load_workbook(
            filename=os.path.join(folder, excel_file))
        self.worksheet = self.workbook.worksheets[sheet_no]

    def clear(self):
        self.worksheet = None
        self.workbook = None

    # get row range
    def _calculate_row_range_(self):
        bounds = sorted([
            merged_range.bounds
            for merged_range in self.worksheet.merged_cells.ranges
            if merged_range.bounds[0] == 1
        ])
        row_block_bound = bounds[self.block_no]
        self.start_row = row_block_bound[1]
        self.end_row = row_block_bound[3]

        print(
            f"source block {self.block_no} row range: A{self.start_row}:A{self.end_row}"
        )

    def calculate_column_range(self):
        self.start_column = 1
        self.end_column = 0
        while type(self.worksheet.cell(
                row=1, column=self.end_column +
                1)).__name__ == 'MergedCell' or self.worksheet.cell(
                    row=1, column=self.end_column + 1).value is not None:
            self.end_column = self.end_column + 1

        print(
            f'sheet {self.worksheet.title} column range: {self.start_column}:{self.end_column}'
        )

    def copy_excel_block(self):
        self._calculate_row_range_()
        self.calculate_column_range()

        copied_range = copy_range(self.start_column, self.start_row,
                                  self.end_column, self.end_row,
                                  self.worksheet)
        return self.start_column, self.start_row, self.end_column, self.end_row, copied_range

    def get_merged_cell_range(self):
        block_area = CellRange(min_col=self.start_column,
                               max_col=self.end_column,
                               min_row=self.start_row,
                               max_row=self.end_row)
        return self.start_row, block_area, self.worksheet.merged_cell_ranges

    def get_column_dimensions(self):
        return getattr(self.worksheet, 'column_dimensions')

    def get_row_dimensions(self):
        return self.worksheet.row_dimensions, self.start_row
