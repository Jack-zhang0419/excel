from openpyxl import load_workbook, workbook
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.merge import MergedCellRange
from util.excel_util import paste_range
from combine_configure import *


class TargetExcel(object):
    """docstring for TargetExcel."""
    def __init__(self, saved_file_path):
        self.saved_file_path = saved_file_path

        self.workbook = workbook.Workbook()
        self.workbook.active.title = SHEET_NAME[0]
        self.workbook.create_sheet(title=SHEET_NAME[1])

        # because two sheets, each one start from row 1
        self.start_rows = [1, 1]
        self.start_column = 1
        self.block_nos = [0, 0]

    def save(self):
        self.workbook.save(self.saved_file_path)

    def set_start_row(self):
        self.start_rows[self.sheet_no] = self.end_row + 1

    def increase_block_no(self):
        self.block_nos[self.sheet_no] += 1

    def switch_sheet(self, sheet_no: int):
        self.sheet_no = sheet_no
        self.worksheet = self.workbook.worksheets[self.sheet_no]

    def paste_excel_block(self, start_column, start_row, end_column, end_row,
                          copied_range):
        self.end_row = self.start_rows[self.sheet_no] + end_row - start_row
        paste_range(start_column, self.start_rows[self.sheet_no], end_column,
                    self.end_row, self.worksheet, copied_range,
                    self.block_nos[self.sheet_no])
        print(
            f"target block {self.block_nos[self.sheet_no]} row range: A{self.start_rows[self.sheet_no]}:A{self.end_row}"
        )

    # https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/worksheet/cell_range.html?highlight=CellRange#
    def paste_merged_cell_range(self, source_start_row: int,
                                source_block_area: CellRange,
                                merged_cell_ranges: list[MergedCellRange]):
        for mcr in merged_cell_ranges:
            if mcr.coord in source_block_area:
                cr = CellRange(mcr.coord)
                cr.shift(row_shift=self.start_rows[self.sheet_no] -
                         source_start_row)
                self.worksheet.merge_cells(cr.coord)

    def set_worksheet_dimensions(self):
        # https://openpyxl.readthedocs.io/en/stable/api/openpyxl.worksheet.dimensions.html#openpyxl.worksheet.dimensions.ColumnDimension.width
        column_widths = SHEET_COLUMN_WIDTH[self.sheet_no]
        for key in column_widths:
            self.worksheet.column_dimensions[key].width = column_widths[key]

        # for attr in ('row_dimensions', 'column_dimensions'):
        #     src = getattr(
        #         self.source_workbook.worksheets[source_sheet_no], attr)
        #     target = getattr(
        #         self.target_workbook.worksheets[target_sheet_no], attr)
        #     for key, dim in src.items():
        #         target[key] = copy(dim)
        #         target[
        #             key].worksheet = self.target_workbook.worksheets[
        #                 target_sheet_no]
